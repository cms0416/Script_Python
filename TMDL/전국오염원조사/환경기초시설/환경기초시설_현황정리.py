# %% [markdown]
# # 환경기초시설 현황 정리 및 유역 확인

# %% [markdown]
# ******************************************************************************
# ### 1. 시설 현황 확인

# %% ---------------------------------------------------------------------------
# 1-1. 라이브러리 로드 및 환경 설정
import pandas as pd
import numpy as np
from pathlib import Path
import warnings
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# 경고 메시지 숨김
warnings.filterwarnings('ignore')

# 경로 설정
BASE_DIR = Path(r"C:/Coding/TMDL/전국오염원조사")
DATA_DIR = BASE_DIR / "기준배출수질" / "환경기초시설 데이터"
STP_DIR = BASE_DIR / "환경기초시설" / "환경기초시설 현황"
OUTPUT_FILE = STP_DIR / "환경기초시설_현황.xlsx"
WATERSHED_FILE = BASE_DIR / "동리별_유역현황.xlsx"

# 디렉토리 생성
STP_DIR.mkdir(parents=True, exist_ok=True)

print(f"작업 경로: {BASE_DIR.absolute()}")

# %% ---------------------------------------------------------------------------
# 1-2. 데이터 불러오기 및 정리

# 환경기초시설 데이터 목록 확인 (xls, xlsx 모두 포함)
files = list(DATA_DIR.glob("*.xls*"))

# map_dfr 대체: 리스트 내포와 concat을 이용한 병합
dfs = []
for f in files:
    # skip=3은 pandas에서 skiprows=3으로 대응하며, col_names=F는 header=None으로 처리
    df = pd.read_excel(f, sheet_name="환경기초시설", skiprows=3, header=None)
    dfs.append(df)

if dfs:
    시설현황_원본 = pd.concat(dfs, ignore_index=True)
else:
    시설현황_원본 = pd.DataFrame()
    print("환경기초시설 원본 데이터를 찾을 수 없습니다.")

# 기존 환경기초시설별 단위유역 현황 자료 불러오기
try:
    stp_유역 = pd.read_excel(OUTPUT_FILE)
    stp_유역 = stp_유역[['시설코드', '단위유역']]
except FileNotFoundError:
    stp_유역 = pd.DataFrame(columns=['시설코드', '단위유역'])
    print("기존 환경기초시설_현황.xlsx 파일을 찾을 수 없어 빈 데이터프레임으로 처리합니다.")

# 데이터 정리 및 가공
if not 시설현황_원본.empty:
    # R: 2:10, 18:20, 26 (1-based index) -> Python: 1:9, 17:19, 25 (0-based index)
    cols_to_select = [1, 2, 3, 4, 5, 6, 7, 8, 9, 17, 18, 19, 25]
    
    시설현황_정리 = 시설현황_원본.iloc[:, cols_to_select].copy()
    시설현황_정리.columns = [
        "시설명", "시설코드", "구분", "시도", "시군", "읍면동", "리",
        "본번", "부번", "물리", "생물", "고도", "가동개시"
    ]

    # 문자열 결합을 위한 주소 및 동리코드 생성 함수 정의
    def make_address(row):
        리 = row['리'] if pd.notna(row['리']) else ""
        본번 = str(row['본번']) if pd.notna(row['본번']) else ""
        부번 = row['부번']
        
        addr = f"강원특별자치도 {row['시군']} {row['읍면동']} "
        if 리: 
            addr += f"{리} "
        addr += 본번
        
        # 부번이 결측치가 아니며 0이 아닌 경우에만 하이픈 추가
        if pd.notna(부번) and str(부번).strip() not in ["0", "0.0", ""]:
            try:
                부번_str = str(int(float(부번)))
            except ValueError:
                부번_str = str(부번)
            addr += f"-{부번_str}"
        return addr.strip()

    def make_dongri_code(row):
        리 = row['리'] if pd.notna(row['리']) else ""
        code = f"{row['시군']} {row['읍면동']} {리}".strip()
        # 다중 공백 제거
        return " ".join(code.split())

    시설현황_정리['주소'] = 시설현황_정리.apply(make_address, axis=1)
    시설현황_정리['동리코드'] = 시설현황_정리.apply(make_dongri_code, axis=1)

    # 문자형 수치 데이터를 숫자로 변환 및 NA를 0으로 대체
    for col in ["물리", "생물", "고도"]:
        시설현황_정리[col] = pd.to_numeric(시설현황_정리[col], errors='coerce').fillna(0)
    
    시설현황_정리['용량'] = 시설현황_정리['물리'] + 시설현황_정리['생물'] + 시설현황_정리['고도']

    # 가동개시 년도 추출 (앞 4자리)
    시설현황_정리['가동개시'] = 시설현황_정리['가동개시'].astype(str).str[:4]
    시설현황_정리['가동개시'] = pd.to_numeric(시설현황_정리['가동개시'], errors='coerce')

    # 시설 종류 분류 (시설코드 7번째 자리, 인덱스 6)
    def get_type(code):
        if pd.isna(code) or len(str(code)) < 7: 
            return np.nan
        c = str(code)[6]
        mapping = {
            'W': '하수처리시설', 'V': '마을하수도', 'A': '농공단지폐수처리시설',
            'I': '산업단지폐수처리시설', 'F': '분뇨처리시설', 'S': '축산폐수처리시설',
            'H': '오수처리시설'
        }
        return mapping.get(c, np.nan)

    시설현황_정리['종류'] = 시설현황_정리['시설코드'].apply(get_type)

    # 열 순서 재배치 (relocate)
    new_cols = [
        "시설명", "시설코드", "종류", "구분", "시도", "시군", "읍면동", "리",
        "본번", "부번", "용량", "물리", "생물", "고도", "가동개시", "주소", "동리코드"
    ]
    시설현황_정리 = 시설현황_정리[new_cols]

    # 단위유역 자료 left_join 및 중복 제거
    시설현황_정리 = 시설현황_정리.merge(stp_유역, on="시설코드", how="left")
    시설현황_정리 = 시설현황_정리.drop_duplicates().reset_index(drop=True)
    
    
print("시설현황 자료 정리 완료. 크기:", 시설현황_정리.shape)
display(시설현황_정리.head(10))


# %% [markdown]
# ******************************************************************************
# ### 2. 단위유역 미확인 시설 유역 확인

# 단위유역이 결측치인 시설에 대해 동리별 유역현황 자료를 이용해 유역 확인
if not 시설현황_원본.empty:
    if WATERSHED_FILE.exists():
        동리별_유역현황 = pd.read_excel(WATERSHED_FILE)
        
        # 유역 점유율이 100%인 지역만 선택
        동리별_유역현황 = 동리별_유역현황[동리별_유역현황['점유율'] == 100]
        동리별_유역현황 = 동리별_유역현황[['동리코드', '단위유역']].rename(columns={'단위유역': '동리별_유역'})

        # 시설현황과 동리별 유역현황 매칭 (left_join)
        시설현황_정리 = 시설현황_정리.merge(동리별_유역현황, on="동리코드", how="left")

        # 단위유역이 결측치(NA)일 경우 동리별_유역 값으로 대체
        시설현황_정리['단위유역'] = 시설현황_정리['단위유역'].fillna(시설현황_정리['동리별_유역'])
        
        # 유역 확인용 임시 열 삭제
        시설현황_정리 = 시설현황_정리.drop(columns=['동리별_유역'])
    else:
        print(f"동리별_유역현황 파일을 찾을 수 없습니다: {WATERSHED_FILE}")


# %% [markdown]
# ******************************************************************************
# ### 3. 정리자료 내보내기

if not 시설현황_원본.empty:
    # Pandas의 ExcelWriter를 openpyxl 엔진으로 실행
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        시설현황_정리.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        
        # 서식 객체 정의
        thin_border = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        header_font = Font(bold=True)
        
        # 1. 셀 테두리, 2. 가운데 정렬, 3. 첫 행 텍스트 볼드 처리
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
                if cell.row == 1:
                    cell.font = header_font
                    
        # 4. 데이터 길이에 맞춰 열 너비 지정
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                try:
                    # 셀 데이터의 문자열 길이 계산 (한글과 영문/숫자 폭 차이를 대략적으로 보정)
                    cell_value = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1.2 for c in cell_value)
                    if length > max_length:
                        max_length = length
                except:
                    pass
            
            # 최소 너비를 10으로 설정하고 여유 공간 확보
            adjusted_width = max(max_length + 2, 10)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"서식 지정 및 파일 저장 완료: {OUTPUT_FILE}")
# %%
